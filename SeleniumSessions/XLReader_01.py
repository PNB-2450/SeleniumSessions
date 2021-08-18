#import xlrd
import openpyxl


path = "D:\Testdata.xlsx"

wb = openpyxl.load_workbook(path)
sheet = wb.active
data = sheet.cell(row=1,column=1)
print(data.value)

rowCount = sheet.max_row
colCount = sheet.max_column

print("Total rows are ", rowCount," : ",'Toatal columns are ', colCount)

for data in range(10):
    print(data.value)


"""
workbook = xlrd.open_workbook(path)
sheet = workbook.sheet_by_name("Data")
rowCount = sheet.nrows
colCount = sheet.ncols
print("Total rows: ", rowCount,' : ', "Totla colonms: ", colCount)

"""